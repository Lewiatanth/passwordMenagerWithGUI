import tkinter as tk
from tkinter import ttk, messagebox
import random
import string
import openpyxl
import os


def check_file():
    if not os.path.exists("data.xlsx"):
        wb = openpyxl.Workbook()
        ws = wb.active

        ws["A2"] = "Witryna"
        ws["B2"] = "Login"
        ws["C2"] = "Hasło"

        wb.save("data.xlsx")


def check_for_empty_cell():
    wb = openpyxl.load_workbook("data.xlsx")
    ws = wb.active
    row = 1
    while ws.cell(row=row, column=1).value is not None:
        row += 1
    return row


def reset_spinbox_value(spinbox):
    spinbox.configure(state="normal")
    spinbox.delete(0, "end")
    spinbox.insert(1, "1")


class PasswordManager:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title("Menadżer haseł")
        photo = tk.PhotoImage(file="key.png")
        self.root.iconphoto(False, photo)

        style = ttk.Style(self.root)
        self.root.tk.call("source", "breeze-dark.tcl")
        style.theme_use("breeze-dark")

        w = 500
        h = 350
        ws = self.root.winfo_screenwidth()
        hs = self.root.winfo_screenheight()
        x = (ws / 2) - (w / 2)
        y = (hs / 2) - (h / 2)
        self.root.geometry('+%d+%d' % (x, y))

        self.font_large = ("Arial", 20)
        self.font_medium = ("Arial", 18)
        self.font_small = ("Arial", 14)
        self.font_size_ten = ("Arial", 10)
        style.configure("yes/no.TButton", font=("Arial", 12))

        self.password_length = 0
        self.lowercase_quantity = 0
        self.uppercase_quantity = 0
        self.digits_quantity = 0
        self.special_quantity = 0
        self.free_characters = 0
        self.final_password = ""
        self.password = []

        self.create_frames()
        self.create_widgets()
        check_file()
        self.set_app_password()

    def create_frames(self):
        self.frame_main = tk.Frame(self.root)
        self.frame_generate_main = tk.Frame(self.root)
        self.frame_view_main = tk.Frame(self.root)
        self.app_password_frame = tk.Frame(self.root)
        self.frame_view_password_check = ttk.Frame(self.frame_view_main)
        self.frame_view_passwords_list = ttk.Frame(self.frame_view_main)

    def create_widgets(self):
        self.welcome_label = ttk.Label(self.frame_main, text="Witaj w menadżerze haseł!", font=self.font_large)
        self.welcome_label.grid(row=0, column=0, columnspan=3, sticky="nsew", padx=10, pady=10)

        self.choice_label = ttk.Label(self.frame_main, text="Wybierz co chcesz zrobić:", font=self.font_medium)
        self.choice_label.grid(row=1, column=0, columnspan=3, pady=10)

        self.button_generate = ttk.Button(self.frame_main, text="Wygeneruj hasło",
                                          command=lambda: self.show_generate_frame())
        self.button_generate.grid(row=2, column=0, sticky="nsew", padx=15, pady=5)

        self.button_view = ttk.Button(self.frame_main, text="Odczytaj hasła", command=lambda: self.show_view_frame())
        self.button_view.grid(row=2, column=1, sticky="nsew", padx=15, pady=5)

        # -----------------------------GenerateFrame-----------------------------------------------------
        self.password_length_label = ttk.Label(self.frame_generate_main, text="Podaj długość hasła (minimum 8 znaków)",
                                               font=self.font_small)
        self.lowercase_quantity_label = ttk.Label(self.frame_generate_main, text="Podaj ilość małych liter (minimum 1)",
                                                  font=self.font_small)
        self.uppercase_quantity_label = ttk.Label(self.frame_generate_main,
                                                  text="Podaj ilość wielkich liter (minimum 1)",
                                                  font=self.font_small)
        self.digits_quantity_label = ttk.Label(self.frame_generate_main, text="Podaj ilość cyfr (minimum 1)",
                                               font=self.font_small)
        self.special_quantity_label = ttk.Label(self.frame_generate_main,
                                                text="Podaj ilość znaków specjlanych (minimum 1)",
                                                font=self.font_small)
        self.free_characters_label = ttk.Label(self.frame_generate_main, font=("Arial", 8))
        self.final_password_label = ttk.Label(self.frame_generate_main, font=self.font_small)
        self.save_question_label = ttk.Label(self.frame_generate_main, text="Czy chcesz zapisać hasło?",
                                             font=("Arial", 8))
        self.done_label = ttk.Label(self.frame_generate_main, text="Hasło nie zostanie zapisane.",
                                    font=self.font_size_ten)
        self.website_label = ttk.Label(self.frame_generate_main, text="Podaj adres witryny:", font=self.font_size_ten)
        self.login_label = ttk.Label(self.frame_generate_main, text="Podaj login:", font=self.font_size_ten)
        self.password_saved_label = ttk.Label(self.frame_generate_main, text="Hasło zostało zapisane!",
                                              font=self.font_size_ten)

        self.password_length_spinbox = ttk.Spinbox(self.frame_generate_main)
        self.lowercase_quantity_spinbox = ttk.Spinbox(self.frame_generate_main)
        self.uppercase_quantity_spinbox = ttk.Spinbox(self.frame_generate_main)
        self.digits_quantity_spinbox = ttk.Spinbox(self.frame_generate_main)
        self.special_quantity_spinbox = ttk.Spinbox(self.frame_generate_main)

        self.ok_length_button = ttk.Button(self.frame_generate_main, text="OK",
                                           command=lambda: self.get_password_length())
        self.ok_lowercase_button = ttk.Button(self.frame_generate_main, text="OK",
                                              command=lambda: self.get_lowercase_quantity())
        self.ok_uppercase_button = ttk.Button(self.frame_generate_main, text="OK",
                                              command=lambda: self.get_uppercase_quantity())
        self.ok_digits_button = ttk.Button(self.frame_generate_main, text="OK",
                                           command=lambda: self.get_digits_quantity())
        self.ok_special_button = ttk.Button(self.frame_generate_main, text="OK",
                                            command=lambda: self.get_special_quantity())
        self.generate_button = ttk.Button(self.frame_generate_main, text="Generuj hasło",
                                          command=lambda: self.generate_password())
        self.save_yes_button = ttk.Button(self.frame_generate_main, text="Tak", style="yes/no.TButton",
                                          command=lambda: self.save_new_password_widgets())
        self.save_no_button = ttk.Button(self.frame_generate_main, text="Nie", style="yes/no.TButton",
                                         command=lambda: self.done_message())
        self.ok_save_password_button = ttk.Button(self.frame_generate_main, text="OK",
                                                  command=lambda: self.save_new_password())
        self.back_button_generate = ttk.Button(self.frame_generate_main, text="Powrót",
                                               command=lambda: self.show_main_frame(self.frame_generate_main))

        self.website_entry = ttk.Entry(self.frame_generate_main, font=self.font_size_ten)

        self.login_entry = ttk.Entry(self.frame_generate_main, font=self.font_size_ten)

        # -------------------------------------View Frame--------------------------------------------------

        self.password_entry_label = ttk.Label(self.frame_view_password_check,
                                              text="Aby odczytać zapisane dane wpisz hasło:",
                                              font=self.font_small)

        self.password_accept_button = ttk.Button(self.frame_view_password_check, text="OK",
                                                 command=lambda: self.check_app_password())
        self.back_button_view = ttk.Button(self.frame_view_password_check, text="Powrót",
                                           command=lambda: self.show_main_frame(self.frame_view_main))

        self.password_entry = ttk.Entry(self.frame_view_password_check, show="*", font=("Arial", 8))
        self.password_entry.bind("<Return>", lambda event: self.check_app_password())

        self.cols = ("Witryna", "Login", "Hasło")

        self.treeview = ttk.Treeview(self.frame_view_passwords_list, show="headings", columns=self.cols, height=10)
        self.treeview.column("Witryna", width=100)
        self.treeview.column("Login", width=100)
        self.treeview.column("Hasło", width=100)

        # --------------------------------Set App Password---------------------------------------------------
        self.app_password_label_one = ttk.Label(self.app_password_frame,
                                                text="Podczas pierwszego uruchomienia ustaw hasło służące do odczytania zapisanych danych:",
                                                wraplength=250, font=self.font_size_ten)
        self.app_password_saved_label = ttk.Label(self.app_password_frame, text="Hasło zostało ustawione!",
                                                  font=self.font_small)

        self.app_passwrod_button = ttk.Button(self.app_password_frame, text="OK",
                                              command=lambda: self.save_app_password())
        self.app_password_ok_button = ttk.Button(self.app_password_frame, text="OK",
                                                 command=lambda: self.show_main_frame(self.app_password_frame))

        self.app_password_entry_one = ttk.Entry(self.app_password_frame, show="*", font=("Arial", 8))
        self.app_password_entry_two = ttk.Entry(self.app_password_frame, show="*", font=("Arial", 8))

    def show_generate_frame(self):
        self.frame_main.pack_forget()
        self.frame_generate_main.pack()
        self.password_length_label.grid(row=0, column=0, columnspan=3, padx=15, pady=5, sticky="nsew")
        self.password_length_spinbox.configure(state="normal")
        self.password_length_spinbox.delete(0, "end")
        self.password_length_spinbox.insert(8, "8")
        self.password_length_spinbox.configure(state="readonly", from_=8, to=100)
        self.ok_length_button.configure(state="normal")
        self.ok_lowercase_button.configure(state="normal")
        self.ok_uppercase_button.configure(state="normal")
        self.ok_digits_button.configure(state="normal")
        self.ok_special_button.configure(state="normal")
        self.password_length_spinbox.grid(row=1, column=0, padx=15, pady=5, sticky="nsew")
        self.ok_length_button.grid(row=1, column=1, columnspan=2, padx=15, pady=5, sticky="nsew")
        self.back_button_generate.grid(row=17, column=1, padx=15, pady=(10, 5))

    def show_view_frame(self):
        self.password_entry.configure(state="normal")
        self.password_accept_button.configure(state="normal")
        self.treeview.delete(*self.treeview.get_children())
        self.frame_main.pack_forget()
        self.frame_view_main.pack()
        self.frame_view_password_check.grid(row=0, column=0)
        self.frame_view_passwords_list.grid(row=0, column=1)
        self.password_entry_label.grid(row=0, column=0, padx=15, pady=5, sticky="w")
        self.password_entry.grid(row=1, column=0, padx=15, pady=5)
        self.password_accept_button.grid(row=2, column=0, padx=15, pady=5)
        self.back_button_view.grid(row=3, column=0, padx=15, pady=5)
        self.treeview.pack()
        wb = openpyxl.load_workbook("data.xlsx")
        ws = wb.active
        list_values = list(ws.values)
        for col_name in list_values[1]:
            self.treeview.heading(col_name, text=col_name)

    def show_main_frame(self, frame):
        for widget in frame.winfo_children():
            widget.grid_forget()
        frame.pack_forget()
        self.frame_main.pack()

    def get_password_length(self):
        self.free_characters = self.password_length = int(self.password_length_spinbox.get())
        if self.password_length >= 8:
            self.password_length_spinbox.configure(state="disable")
            self.ok_length_button.configure(state="disable")
            self.lowercase_quantity_label.grid(row=2, column=0, columnspan=3, padx=15, pady=5, sticky="nsew")
            reset_spinbox_value(self.lowercase_quantity_spinbox)
            self.lowercase_quantity_spinbox.configure(state="readonly", from_=1, to=self.free_characters - 3)
            self.lowercase_quantity_spinbox.grid(row=3, column=0, padx=15, pady=5, sticky="nsew")
            self.ok_lowercase_button.grid(row=3, column=1, columnspan=2, padx=15, pady=5, sticky="nsew")

    def get_lowercase_quantity(self):
        self.lowercase_quantity = int(self.lowercase_quantity_spinbox.get())
        if 1 <= self.lowercase_quantity <= self.free_characters - 3:
            self.free_characters -= self.lowercase_quantity
            self.lowercase_quantity_spinbox.configure(state="disable")
            self.ok_lowercase_button.configure(state="disable")
            self.uppercase_quantity_label.grid(row=4, column=0, columnspan=3, padx=15, pady=5, sticky="nsew")
            reset_spinbox_value(self.uppercase_quantity_spinbox)
            self.uppercase_quantity_spinbox.configure(state="readonly", from_=1, to=self.free_characters - 2)
            self.uppercase_quantity_spinbox.grid(row=5, column=0, padx=15, pady=5, sticky="nsew")
            self.ok_uppercase_button.grid(row=5, column=1, columnspan=2, padx=15, pady=5, sticky="nsew")

    def get_uppercase_quantity(self):
        self.uppercase_quantity = int(self.uppercase_quantity_spinbox.get())
        if 1 <= self.uppercase_quantity <= self.free_characters - 2:
            self.free_characters -= self.uppercase_quantity
            self.uppercase_quantity_spinbox.configure(state="disable")
            self.ok_uppercase_button.configure(state="disable")
            self.digits_quantity_label.grid(row=6, column=0, columnspan=3, padx=15, pady=5, sticky="nsew")
            reset_spinbox_value(self.digits_quantity_spinbox)
            self.digits_quantity_spinbox.configure(state="readonly", from_=1, to=self.free_characters - 1)
            self.digits_quantity_spinbox.grid(row=7, column=0, padx=15, pady=5, sticky="nsew")
            self.ok_digits_button.grid(row=7, column=1, columnspan=2, padx=15, pady=5, sticky="nsew")

    def get_digits_quantity(self):
        self.digits_quantity = int(self.digits_quantity_spinbox.get())
        if 1 <= self.digits_quantity <= self.free_characters - 1:
            self.free_characters -= self.digits_quantity
            self.digits_quantity_spinbox.configure(state="disable")
            self.ok_digits_button.configure(state="disable")
            self.special_quantity_label.grid(row=8, column=0, columnspan=3, padx=15, pady=5, sticky="nsew")
            reset_spinbox_value(self.special_quantity_spinbox)
            self.special_quantity_spinbox.configure(state="readonly", from_=1, to=self.free_characters)
            self.special_quantity_spinbox.grid(row=9, column=0, padx=15, pady=5, sticky="nsew")
            self.ok_special_button.grid(row=9, column=1, columnspan=2, padx=15, pady=5, sticky="nsew")

    def get_special_quantity(self):
        self.special_quantity = int(self.special_quantity_spinbox.get())
        if 1 <= self.special_quantity <= self.free_characters:
            if self.free_characters > 1:
                self.free_characters -= self.special_quantity
                if self.free_characters > 0:
                    self.free_characters_label.configure(
                        text=f"Pozostało {self.free_characters} wolnych znaków, zostaną uzupełnione małymi literami."
                    )
                    self.free_characters_label.grid(row=10, column=0, columnspan=3, padx=15, pady=5, sticky="nsew")
                    self.lowercase_quantity += self.free_characters
            self.generate_button.grid(row=11, column=0, padx=15, pady=5, sticky="nsew")
            self.special_quantity_spinbox.configure(state="disable")
            self.ok_special_button.configure(state="disable")

    def generate_password(self):
        lower = self.lowercase_quantity
        upper = self.uppercase_quantity
        special = self.special_quantity
        digits = self.digits_quantity
        self.password.clear()
        for i in range(self.password_length):
            if lower > 0:
                self.password.append(random.choice(string.ascii_lowercase))
                lower -= 1
            if upper > 0:
                self.password.append(random.choice(string.ascii_uppercase))
                upper -= 1
            if special > 0:
                self.password.append(random.choice(string.punctuation))
                special -= 1
            if digits > 0:
                self.password.append(random.choice(string.digits))
                digits -= 1
        random.shuffle(self.password)
        self.final_password = "".join(self.password)
        self.final_password_label.configure(text=f"Twoje hasło to: {self.final_password}")
        self.final_password_label.grid(row=12, column=0, columnspan=3, padx=15, pady=5, sticky="nsew")
        self.save_question_label.grid(row=13, column=0, columnspan=3, padx=15, pady=5, sticky="nsew")
        self.save_yes_button.grid(row=14, column=0, padx=15, pady=5, ipadx=15, ipady=5, sticky="nsew")
        self.save_no_button.grid(row=14, column=1, columnspan=2, padx=15, pady=5, ipadx=15, ipady=5, sticky="nsew")

    def set_app_password(self):
        wb = openpyxl.load_workbook("data.xlsx")
        ws = wb.active
        if ws["A1"].value is None:
            self.app_password_frame.pack()
            self.app_password_label_one.grid(row=0, column=0, columnspan=2, padx=10, pady=5)
            self.app_password_entry_one.grid(row=1, column=0, padx=(5, 0), pady=5, sticky="w")
            self.app_password_entry_two.grid(row=2, column=0, padx=(5, 0), pady=5, sticky="w")
            self.app_passwrod_button.grid(row=1, rowspan=2, column=1, padx=(0, 5), pady=5, sticky="w")

        else:
            self.frame_main.pack()

    def save_app_password(self):
        password_one = self.app_password_entry_one.get()
        password_two = self.app_password_entry_two.get()
        if password_one == password_two and len(password_one) > 0:
            wb = openpyxl.load_workbook("data.xlsx")
            ws = wb.active
            ws["A1"] = password_one
            wb.save("data.xlsx")
            for widget in self.app_password_frame.winfo_children():
                widget.grid_forget()
            self.app_password_saved_label.pack(padx=5, pady=10)
            self.app_password_ok_button.pack(padx=5, pady=10)
        else:
            messagebox.showerror(title="Błąd!", message="Podane hasła różnią się. Wpisz je jeszcze raz.")

    def check_passwords_correctness(self):
        if str(self.app_password_entry_one.get()) == str(self.app_password_entry_two.get()):
            wb = openpyxl.load_workbook("data.xlsx")
            ws = wb.active
            ws["A1"] = self.app_password_entry_one.get()
            wb.save("data.xlsx")
            messagebox.showinfo(message="Hasło ustawione!")
        else:
            messagebox.showerror(message="Podane hasła różnią się!")

    def check_app_password(self):
        wb = openpyxl.load_workbook("data.xlsx")
        ws = wb.active
        answer = self.password_entry.get()
        if answer == str(ws.cell(row=1, column=1).value):
            self.password_entry.delete(0, "end")
            self.password_entry.configure(state="disable")
            self.password_accept_button.configure(state="disable")
            list_values = list(ws.values)
            for value_tuple in list_values[2:]:
                self.treeview.insert("", tk.END, values=value_tuple)
        else:
            self.password_entry.delete(0, "end")
            messagebox.showerror(title="Błędne hasło", message="Hasło nieprawidłowe")

    def done_message(self):
        self.save_question_label.grid_forget()
        self.save_yes_button.grid_forget()
        self.save_no_button.grid_forget()
        self.done_label.grid(row=13, column=0, columnspan=2, padx=15, pady=5)

    def save_new_password_widgets(self):
        self.save_question_label.grid_forget()
        self.save_yes_button.grid_forget()
        self.save_no_button.grid_forget()
        self.website_label.grid(row=13, column=0, padx=15, pady=5)
        self.login_label.grid(row=13, column=1, padx=15, pady=5)
        self.website_entry.grid(row=14, column=0, padx=15, pady=5)
        self.login_entry.grid(row=14, column=1, padx=15, pady=5)
        self.ok_save_password_button.grid(row=15, column=0, padx=15, pady=5)

    def save_new_password(self):
        website = self.website_entry.get()
        login = self.login_entry.get()
        wb = openpyxl.load_workbook("data.xlsx")
        ws = wb.active
        row = check_for_empty_cell()
        row_string = str(row)
        site_row = "A" + row_string
        login_row = "B" + row_string
        password_row = "C" + row_string
        ws[site_row] = website
        ws[login_row] = login
        ws[password_row] = self.final_password
        wb.save("data.xlsx")
        self.website_label.grid_forget()
        self.website_entry.grid_forget()
        self.login_label.grid_forget()
        self.login_entry.grid_forget()
        self.ok_save_password_button.grid_forget()
        self.password_saved_label.grid(row=13, column=0, columnspan=2, padx=15, pady=5)

    def run(self):
        self.root.mainloop()


if __name__ == "__main__":
    password_manager = PasswordManager()
    password_manager.run()
